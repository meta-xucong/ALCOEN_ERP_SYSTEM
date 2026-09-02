"""Statement generation and listing regression tests."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from app.services.contract_service import ContractService


def _create_contract_with_transaction(
    owner_user_id: int,
    company_name: str = "Statement Corp",
    product_type: str = "TypeS",
    delivery_date: str = "2026-04-01",
    product_code: str = "ST-P1",
):
    """Create one contract and one transaction for statement generation."""
    contract = ContractService.create_contract(
        {
            "contract_no": f"ST-{owner_user_id}-{company_name}",
            "company_name": company_name,
            "owner": "Sales - Owner",
            "department": "Sales",
            "manager": "Sales Owner",
            "created_by_id": owner_user_id,
        },
        [
            {
                "product_code": product_code,
                "product_name": "Statement Product",
                "product_model": "SP1",
                "product_type": product_type,
                "quantity": 10,
                "unit": "pcs",
                "price": 10,
                "remark": "",
            }
        ],
    )
    cp = contract.contract_products[0]
    ContractService.add_transaction(
        contract.id,
        {
            "contract_product_id": cp.id,
            "quantity": 5,
            "unit": "pcs",
            "price_with_tax": 10,
            "handler": "Logistics",
            "delivery_date": delivery_date,
            "invoice_date": "",
            "remark": "tx for statement",
        },
        is_new=True,
    )
    return contract


def test_statement_generator_empty_filter_does_not_crash(app, client, login, base_data):
    """Submitting generator without filters should return page with warning, not 500."""
    login(base_data["superadmin_id"])
    resp = client.post("/statement/generator", data={}, follow_redirects=True)
    assert resp.status_code == 200
    assert b"<form" in resp.data


def test_statement_generator_creates_statement(app, client, login, base_data):
    """Generator should create a statement and redirect to statement detail."""
    from app.models import Statement

    with app.app_context():
        _create_contract_with_transaction(base_data["owner_user_id"])

    login(base_data["superadmin_id"])
    resp = client.post(
        "/statement/generator",
        data={"company_name": "Statement Corp"},
        follow_redirects=False,
    )

    assert resp.status_code == 302
    assert "/statement/DZ" in resp.headers.get("Location", "")

    with app.app_context():
        statement = Statement.query.first()
        assert statement is not None
        assert statement.company_name == "Statement Corp"
        assert statement.record_count == 1


def test_statement_generator_aggregates_multiple_company_cards(app, client, login, base_data):
    """Selected company cards should create one statement with all matching transactions."""
    from app.models import Statement
    import json

    with app.app_context():
        _create_contract_with_transaction(base_data["owner_user_id"], "Statement Corp A")
        _create_contract_with_transaction(base_data["owner_user_id"], "Statement Corp B")

    login(base_data["superadmin_id"])
    generator_page = client.get("/statement/generator")
    assert generator_page.status_code == 200
    assert b'id="companyCards"' in generator_page.data
    assert b'id="addCompanyCard"' in generator_page.data
    assert b'name="company_names"' in generator_page.data

    response = client.post(
        "/statement/generator",
        data={"company_names": ["Statement Corp A", "Statement Corp B", "Statement Corp A"]},
        follow_redirects=False,
    )

    assert response.status_code == 302
    with app.app_context():
        statement = Statement.query.one()
        assert statement.company_name == "2家公司"
        assert statement.record_count == 2
        assert json.loads(statement.filter_products)["company_names"] == ["Statement Corp A", "Statement Corp B"]

    result_page = client.get(response.headers["Location"])
    assert result_page.status_code == 200
    result_html = result_page.get_data(as_text=True)
    assert "Statement Corp A、Statement Corp B" in result_html


def test_statement_generator_product_code_filter_supports_fuzzy_match(app, client, login, base_data):
    """Product code filter should use fuzzy matching like product name."""
    from app.models import Statement

    with app.app_context():
        _create_contract_with_transaction(base_data["owner_user_id"])

    login(base_data["superadmin_id"])
    resp = client.post(
        "/statement/generator",
        data={"product_code_filter": "ST-"},
        follow_redirects=False,
    )

    assert resp.status_code == 302
    assert "/statement/DZ" in resp.headers.get("Location", "")

    with app.app_context():
        statement = Statement.query.first()
        assert statement is not None
        assert statement.record_count == 1


def test_statement_generator_filters_by_contract_created_date(app, client, login, base_data):
    """Date ranges can use contract creation dates instead of delivery dates."""
    from app import db
    from app.models import Statement
    import json

    with app.app_context():
        created_in_range = _create_contract_with_transaction(
            base_data["owner_user_id"],
            "Created In Range",
            delivery_date="2026-04-01",
        )
        created_out_of_range = _create_contract_with_transaction(
            base_data["owner_user_id"],
            "Created Out Of Range",
            delivery_date="2026-01-15",
        )
        created_in_range.created_at = datetime(2026, 1, 15, 18, 30)
        created_out_of_range.created_at = datetime(2026, 3, 1, 9, 0)
        db.session.commit()

    login(base_data["superadmin_id"])
    response = client.post(
        "/statement/generator",
        data={
            "date_filter_mode": "contract_created_at",
            "start_date": "2026-01-15",
            "end_date": "2026-01-15",
        },
        follow_redirects=False,
    )

    assert response.status_code == 302
    with app.app_context():
        statement = Statement.query.one()
        assert statement.record_count == 1
        assert json.loads(statement.filter_products)["date_filter_mode"] == "contract_created_at"

    result_page = client.get(response.headers["Location"])
    assert result_page.status_code == 200
    result_html = result_page.get_data(as_text=True)
    assert "合同创建日期" in result_html
    assert "2026-01-15" in result_html

    export_page = client.get(f'{response.headers["Location"]}/export')
    assert export_page.status_code == 200
    workbook = load_workbook(BytesIO(export_page.data))
    worksheet = workbook.active
    worksheet_values = [
        cell.value
        for row in worksheet.iter_rows()
        for cell in row
    ]
    assert "合同创建日期" in worksheet_values
    assert "2026-01-15" in worksheet_values


def test_statement_generator_filters_by_product_type_fuzzily(app, client, login, base_data):
    """Product type filters use fuzzy matching against delivery snapshots."""
    from app.models import Statement
    import json

    with app.app_context():
        _create_contract_with_transaction(
            base_data["owner_user_id"],
            "Column Accessory Corp",
            product_type="色谱柱配件",
            product_code="ST-COLUMN-001",
        )
        _create_contract_with_transaction(
            base_data["owner_user_id"],
            "Tube Accessory Corp",
            product_type="柱管配件",
            product_code="ST-TUBE-001",
        )

    login(base_data["superadmin_id"])
    generator_page = client.get("/statement/generator")
    assert generator_page.status_code == 200
    assert b'name="date_filter_mode"' in generator_page.data
    assert b'name="product_type_filter"' in generator_page.data
    assert b'date-filter-mode' in generator_page.data

    response = client.post(
        "/statement/generator",
        data={"product_type_filter": "色谱柱"},
        follow_redirects=False,
    )

    assert response.status_code == 302
    with app.app_context():
        statement = Statement.query.one()
        assert statement.record_count == 1
        assert json.loads(statement.filter_products)["product_types"] == ["色谱柱"]


def test_statement_list_requires_login(client):
    """Statement list should redirect anonymous users to login."""
    resp = client.get("/statement/list", follow_redirects=False)
    assert resp.status_code == 302
    assert "/auth/login" in resp.headers.get("Location", "")


def test_multi_department_pm_generates_statement_for_selected_department(
    app, client, login
):
    """A multi-department PM must generate each statement for one chosen department."""
    import json

    from app import db
    from app.models import Department, Role, User, Statement

    with app.app_context():
        sales = Department(name="Sales")
        ops = Department(name="Ops")
        db.session.add_all([sales, ops])
        db.session.flush()

        pm_role = Role(
            name="Department PM",
            code="department_pm",
            permissions=json.dumps(
                ["statement_create", "statement_view", "statement_export"]
            ),
            level=50,
        )
        db.session.add(pm_role)
        db.session.flush()

        pm = User(
            username="pm_statement_multi",
            password_hash="x",
            real_name="PM Statement Multi",
            role_id=pm_role.id,
            department_id=sales.id,
            is_active=True,
            is_superadmin=False,
            require_password_change=False,
        )
        db.session.add(pm)
        db.session.flush()
        pm.set_departments([sales.id, ops.id])
        db.session.commit()

        for index, department_name in enumerate(("Sales", "Ops"), start=1):
            contract = ContractService.create_contract(
                {
                    "contract_no": f"ST-MULTI-{index}",
                    "company_name": "Multi Department Corp",
                    "owner": "PM Statement Multi",
                    "department": department_name,
                    "manager": "PM Statement Multi",
                    "created_by_id": pm.id,
                },
                [
                    {
                        "product_code": f"ST-MULTI-P{index}",
                        "product_name": f"Statement Product {index}",
                        "quantity": 10,
                        "unit": "pcs",
                        "price": 10,
                    }
                ],
            )
            ContractService.add_transaction(
                contract.id,
                {
                    "contract_product_id": contract.contract_products[0].id,
                    "quantity": 5,
                    "unit": "pcs",
                    "price_with_tax": 10,
                    "handler": "PM Statement Multi",
                    "delivery_date": "2026-04-01",
                    "invoice_date": "",
                    "remark": department_name,
                },
                is_new=True,
            )
        pm_id = pm.id

    login(pm_id)

    generator_response = client.get("/statement/generator")
    assert generator_response.status_code == 200
    assert b'name="department"' in generator_response.data
    assert b'value="Sales"' in generator_response.data
    assert b'value="Ops"' in generator_response.data

    create_response = client.post(
        "/statement/generator",
        data={
            "company_name": "Multi Department Corp",
            "department": "Ops",
        },
        follow_redirects=False,
    )
    assert create_response.status_code == 302

    with app.app_context():
        statement = Statement.query.order_by(Statement.id.desc()).first()
        assert statement is not None
        assert statement.department == "Ops"
        assert statement.record_count == 1
