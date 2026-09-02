from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.properties import PageSetupProperties
from datetime import datetime
from itertools import groupby
import json
import os
from PIL import Image as PILImage


def export_statement_to_excel(statement, transactions, output_path):
    """[v1.3] 导出对账单为Excel文件 - 按合同分组显示
    
    Args:
        statement: Statement 对象
        transactions: Transaction 列表
        output_path: 输出文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "对账单"
    
    # 定义样式
    title_font = Font(size=18, bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True, size=11)
    contract_header_font = Font(bold=True, color="FFFFFF", size=12)
    red_font = Font(color="FF0000", bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    contract_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')  # 合同行蓝色
    subtotal_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')  # 小计行浅蓝
    total_fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')  # 合计行绿色
    
    # 设置列宽 [v1.3] 调整列宽
    ws.column_dimensions['A'].width = 6   # 序号
    ws.column_dimensions['B'].width = 14  # 发货日期
    ws.column_dimensions['C'].width = 12  # 产品编码
    ws.column_dimensions['D'].width = 20  # 产品名称
    ws.column_dimensions['E'].width = 12  # 型号
    ws.column_dimensions['F'].width = 8   # 数量
    ws.column_dimensions['G'].width = 6   # 单位
    ws.column_dimensions['H'].width = 11  # 含税单价
    ws.column_dimensions['I'].width = 12  # 含税金额
    
    current_row = 1
    
    # [v1.4] 标题行：左侧文字，右侧Logo
    logo_path = os.path.join(os.path.dirname(__file__), '..', '..', 'static', 'img', 'logo.png')
    
    # 先设置标题行（A-H列，给logo留出I列）
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = '客户对账单'
    ws[f'A{current_row}'].font = Font(size=16, bold=True, color="1F4E79")
    ws[f'A{current_row}'].alignment = center_align
    
    # 在I列（最右侧）放置logo，与标题平行
    if os.path.exists(logo_path):
        with PILImage.open(logo_path) as img:
            orig_width, orig_height = img.size
        
        logo = XLImage(logo_path)
        # 设置目标高度为16像素（与标题字体大小一致），宽度按比例
        target_height = 16
        ratio = target_height / orig_height
        target_width = int(orig_width * ratio)
        
        logo.width = target_width
        logo.height = target_height
        
        # 将logo放在I1（最右侧列），与标题同行
        ws.add_image(logo, 'I1')
    
    ws.row_dimensions[current_row].height = 22
    current_row += 2
    
    # 对账单信息
    ws[f'A{current_row}'] = f'对账单号：{statement.statement_no}'
    ws[f'A{current_row}'].font = bold_font
    current_row += 1
    
    ws[f'A{current_row}'] = f'客户名称：{statement.company_name}'
    ws[f'A{current_row}'].font = bold_font
    current_row += 1

    # Preserve every selected company in aggregate exports, not only the count.
    filter_conditions = {}
    if statement.filter_products:
        try:
            decoded_filter_conditions = json.loads(statement.filter_products)
            if isinstance(decoded_filter_conditions, dict):
                filter_conditions = decoded_filter_conditions
        except (TypeError, ValueError):
            filter_conditions = {}
    use_contract_created_at = (
        filter_conditions.get('date_filter_mode') == 'contract_created_at'
    )
    statement_date_label = (
        '合同创建日期' if use_contract_created_at else '发货日期'
    )
    company_names = filter_conditions.get('company_names') or []
    if company_names:
        ws[f'A{current_row}'] = f'聚合客户：{"、".join(company_names)}'
        ws[f'A{current_row}'].font = bold_font
        ws[f'A{current_row}'].alignment = left_align
        current_row += 1
    
    if statement.filter_start_date and statement.filter_end_date:
        date_filter_label = (
            '合同创建日期'
            if filter_conditions.get('date_filter_mode') == 'contract_created_at'
            else '发货日期'
        )
        ws[f'A{current_row}'] = (
            f'时间范围（{date_filter_label}）：'
            f'{statement.filter_start_date} 至 {statement.filter_end_date}'
        )
    else:
        ws[f'A{current_row}'] = '时间范围：不限'
    current_row += 1
    
    ws[f'A{current_row}'] = f'生成时间：{statement.created_at.strftime("%Y-%m-%d %H:%M")}'
    current_row += 2
    
    # [v1.3] 按合同分组显示
    # 先按合同分组
    grouped = {}
    for trans in transactions:
        contract_no = trans.contract.contract_no if trans.contract else '无合同关联'
        if contract_no not in grouped:
            grouped[contract_no] = {
                'transactions': [],
                'contract': trans.contract
            }
        grouped[contract_no]['transactions'].append(trans)
    
    global_index = 0
    grand_total = 0
    
    for contract_no, group_data in grouped.items():
        contract = group_data['contract']
        contract_trans = group_data['transactions']
        
        # 合同标题行
        ws.merge_cells(f'A{current_row}:I{current_row}')
        company_name = contract.company_name if contract else '未知客户'
        dept_info = f" [{contract.department or '无部门'} - {contract.manager or '无负责人'}]" if contract else ""
        ws[f'A{current_row}'] = f'合同编号：{contract_no} | {company_name}{dept_info}'
        ws[f'A{current_row}'].font = contract_header_font
        ws[f'A{current_row}'].fill = contract_fill
        ws[f'A{current_row}'].alignment = left_align
        for col in range(1, 10):
            ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1
        
        # 表头
        headers = ['序号', statement_date_label, '产品编码', '产品名称', '型号', '数量', '单位', '含税单价', '含税金额']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = header_fill
        current_row += 1
        
        # 合同小计
        contract_total = 0
        
        # 数据行
        for trans in contract_trans:
            global_index += 1
            contract_total += trans.total_price_with_tax
            
            # 序号
            ws.cell(row=current_row, column=1, value=global_index).border = thin_border
            ws.cell(row=current_row, column=1).alignment = center_align
            
            # The detail date must match the active generator filter mode.
            detail_date = (
                trans.contract.created_at
                if use_contract_created_at and trans.contract
                else trans.delivery_date
            )
            if hasattr(detail_date, 'strftime'):
                detail_date = detail_date.strftime('%Y-%m-%d')
            ws.cell(row=current_row, column=2, value=detail_date).border = thin_border
            ws.cell(row=current_row, column=2).alignment = center_align
            
            # 产品编码
            ws.cell(row=current_row, column=3, value=trans.product_code).border = thin_border
            ws.cell(row=current_row, column=3).alignment = center_align
            ws.cell(row=current_row, column=3).font = Font(size=9)
            
            # 产品名称
            ws.cell(row=current_row, column=4, value=trans.product_name or '-').border = thin_border
            ws.cell(row=current_row, column=4).alignment = left_align
            
            # 型号
            ws.cell(row=current_row, column=5, value=trans.product_model or '-').border = thin_border
            ws.cell(row=current_row, column=5).alignment = center_align
            
            # 数量
            ws.cell(row=current_row, column=6, value=trans.quantity).border = thin_border
            ws.cell(row=current_row, column=6).alignment = center_align
            
            # 单位
            ws.cell(row=current_row, column=7, value=trans.unit).border = thin_border
            ws.cell(row=current_row, column=7).alignment = center_align
            
            # 含税单价
            price_cell = ws.cell(row=current_row, column=8, value=trans.price_with_tax)
            price_cell.border = thin_border
            price_cell.number_format = '#,##0.00'
            price_cell.alignment = right_align
            
            # 含税金额
            total_cell = ws.cell(row=current_row, column=9, value=trans.total_price_with_tax)
            total_cell.border = thin_border
            total_cell.number_format = '#,##0.00'
            total_cell.alignment = right_align
            
            current_row += 1
        
        # 合同小计行
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = f'{contract_no} 小计：'
        ws[f'A{current_row}'].alignment = right_align
        ws[f'A{current_row}'].font = bold_font
        ws[f'A{current_row}'].fill = subtotal_fill
        for col in range(1, 9):
            ws.cell(row=current_row, column=col).border = thin_border
            ws.cell(row=current_row, column=col).fill = subtotal_fill
        
        subtotal_cell = ws[f'I{current_row}']
        subtotal_cell.value = contract_total
        subtotal_cell.font = bold_font
        subtotal_cell.number_format = '#,##0.00'
        subtotal_cell.border = thin_border
        subtotal_cell.fill = subtotal_fill
        subtotal_cell.alignment = right_align
        
        grand_total += contract_total
        current_row += 1
        
        # 空行分隔
        current_row += 1
    
    # 合计行
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = '合计：'
    ws[f'A{current_row}'].alignment = right_align
    ws[f'A{current_row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{current_row}'].fill = total_fill
    for col in range(1, 9):
        ws.cell(row=current_row, column=col).border = thin_border
        ws.cell(row=current_row, column=col).fill = total_fill
    
    total_cell = ws[f'I{current_row}']
    total_cell.value = grand_total
    total_cell.font = Font(bold=True, size=12, color="FFFFFF")
    total_cell.number_format = '#,##0.00'
    total_cell.border = thin_border
    total_cell.fill = total_fill
    total_cell.alignment = right_align
    
    # 保存文件
    wb.save(output_path)
    return output_path



def export_delivery_note_to_excel(contract, transactions, output_path, note_no):
    """[v1.3] 导出发货单为Excel文件
    
    Args:
        contract: Contract 对象
        transactions: Transaction 列表（发货记录）
        output_path: 输出文件路径
        note_no: 发货单编号
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "发货单"
    
    # 定义样式
    title_font = Font(size=18, bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    wrapped_left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='212529', end_color='212529', fill_type='solid')

    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.28
    ws.page_margins.right = 0.28
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.45
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    ws.print_options.horizontalCentered = True
    ws.sheet_view.view = 'pageLayout'
    
    # 设置列宽，确保适配 A4 纸宽
    ws.column_dimensions['A'].width = 5.5   # 序号
    ws.column_dimensions['B'].width = 12    # 产品编码
    ws.column_dimensions['C'].width = 17    # 产品名称
    ws.column_dimensions['D'].width = 11    # 型号
    ws.column_dimensions['E'].width = 7     # 数量
    ws.column_dimensions['F'].width = 5.5   # 单位
    ws.column_dimensions['G'].width = 11    # 发货日期
    ws.column_dimensions['H'].width = 9     # 经手人
    ws.column_dimensions['I'].width = 13    # 备注
    
    current_row = 1
    
    # [v1.4] 标题行：左侧文字，右侧Logo
    logo_path = os.path.join(os.path.dirname(__file__), '..', '..', 'static', 'img', 'logo.png')
    
    # 先设置标题行（A-H列，给logo留出I列）
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = '发货单'
    ws[f'A{current_row}'].font = Font(size=16, bold=True, color="1e3a5f")
    ws[f'A{current_row}'].alignment = center_align
    
    # 在I列（最右侧）放置logo，与标题平行
    if os.path.exists(logo_path):
        with PILImage.open(logo_path) as img:
            orig_width, orig_height = img.size
        
        logo = XLImage(logo_path)
        # 设置目标高度为16像素（与标题字体大小一致），宽度按比例
        target_height = 16
        ratio = target_height / orig_height
        target_width = int(orig_width * ratio)
        
        logo.width = target_width
        logo.height = target_height
        
        # 将logo放在I1（最右侧列），与标题同行
        ws.add_image(logo, 'I1')
    
    ws.row_dimensions[current_row].height = 22
    current_row += 2
    
    # 发货单信息
    ws[f'A{current_row}'] = f'发货单号：{note_no}'
    ws[f'A{current_row}'].font = bold_font
    current_row += 1
    
    ws[f'A{current_row}'] = f'合同编号：{contract.contract_no}'
    ws[f'A{current_row}'].font = bold_font
    current_row += 1
    
    ws[f'A{current_row}'] = f'客户名称：{contract.company_name}'
    ws[f'A{current_row}'].font = bold_font
    current_row += 1
    
    if contract.department or contract.manager:
        dept_info = f'{contract.department or ""} - {contract.manager or ""}'
        ws[f'A{current_row}'] = f'部门/负责人：{dept_info}'
        ws[f'A{current_row}'].font = bold_font
        current_row += 1
    
    ws[f'A{current_row}'] = f'生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}'
    current_row += 2
    
    # 表头
    headers = ['序号', '产品编码', '产品名称', '型号', '数量', '单位', '发货日期', '经手人', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill
    ws.row_dimensions[current_row].height = 24
    current_row += 1
    
    # 数据行
    for idx, trans in enumerate(transactions, 1):
        # 序号
        ws.cell(row=current_row, column=1, value=idx).border = thin_border
        ws.cell(row=current_row, column=1).alignment = center_align
        
        # 产品编码
        ws.cell(row=current_row, column=2, value=trans.product_code).border = thin_border
        ws.cell(row=current_row, column=2).alignment = center_align
        
        # 产品名称
        ws.cell(row=current_row, column=3, value=trans.product_name or '-').border = thin_border
        ws.cell(row=current_row, column=3).alignment = wrapped_left_align
        
        # 型号
        ws.cell(row=current_row, column=4, value=trans.product_model or '-').border = thin_border
        ws.cell(row=current_row, column=4).alignment = wrapped_left_align
        
        # 数量
        ws.cell(row=current_row, column=5, value=trans.quantity).border = thin_border
        ws.cell(row=current_row, column=5).alignment = center_align
        
        # 单位
        ws.cell(row=current_row, column=6, value=trans.unit).border = thin_border
        ws.cell(row=current_row, column=6).alignment = center_align
        
        # 发货日期
        delivery_date = trans.delivery_date
        if hasattr(delivery_date, 'strftime'):
            delivery_date = delivery_date.strftime('%Y-%m-%d')
        ws.cell(row=current_row, column=7, value=delivery_date).border = thin_border
        ws.cell(row=current_row, column=7).alignment = center_align
        
        # 经手人
        ws.cell(row=current_row, column=8, value=trans.handler or '-').border = thin_border
        ws.cell(row=current_row, column=8).alignment = center_align
        
        # 备注
        ws.cell(row=current_row, column=9, value=trans.remark or '-').border = thin_border
        ws.cell(row=current_row, column=9).alignment = wrapped_left_align
        ws.row_dimensions[current_row].height = 24
        
        current_row += 1
    
    # 合计行
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = '合计：'
    ws[f'A{current_row}'].alignment = right_align
    ws[f'A{current_row}'].font = bold_font
    for col in range(1, 5):
        ws.cell(row=current_row, column=col).border = thin_border
    
    total_qty = sum(t.quantity for t in transactions)
    ws.cell(row=current_row, column=5, value=total_qty).border = thin_border
    ws.cell(row=current_row, column=5).alignment = center_align
    ws.cell(row=current_row, column=5).font = bold_font
    
    for col in range(6, 10):
        ws.cell(row=current_row, column=col).border = thin_border

    ws.print_area = f'A1:I{current_row}'
    
    # 保存文件
    wb.save(output_path)
    return output_path
